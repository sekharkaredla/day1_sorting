import openpyxl

options_file = openpyxl.load_workbook('day1_options.xlsx')

options_sheet = options_file.get_sheet_by_name('options')

#print options_sheet


options = {}


max_rows_options = options_sheet.max_row


#print max_rows


for each_row in range(1,max_rows_options+1):
    roll_no = str(options_sheet.cell(row = each_row,column = 1).value).strip()
    options[roll_no] = {}
    for preference in range(2,5):
        options[roll_no][preference-1] = str(options_sheet.cell(row = each_row,column = preference).value).strip().upper()

#print options
#------------------------------------------------------------------------------------------------------------------
Infosys_placed = open('Infosys_placed.txt','w')
Accenture_placed = open('Accenture_placed.txt','w')
Wipro_placed = open('Wipro_placed.txt','w')
Unplaced = open('Unplaced.txt','w')
#------------------------------------------------------------------------------------------------------------------
def read_compant_list(company_file_name):
    FILE = openpyxl.load_workbook(company_file_name)
    sheet = FILE.get_sheet_by_name('Sheet1')
    max_rows = sheet.max_row
    selected_list = []
    for each_row in range(1,max_rows+1):
        selected_list.append(str(sheet.cell(row = each_row,column = 1).value).strip())
    return selected_list
#------------------------------------------------------------------------------------------------------------------
def update_company_file(company_name,roll_no):
    roll_no = str(roll_no)+'\n'
    if company_name == 'INFOSYS':
        Infosys_placed.write(roll_no)
    elif company_name == 'ACCENTURE':
        Accenture_placed.write(roll_no)
    elif company_name == 'WIPRO':
        Wipro_placed.write(roll_no)
    else:
        Unplaced.write(roll_no)

#------------------------------------------------------------------------------------------------------------------
# infosys_file = openpyxl.load_workbook('infosys.xlsx')
#
#
# infosys_sheet = infosys_file.get_sheet_by_name('Sheet1')
#
#
# max_rows_infosys = infosys_sheet.max_row
#
#
# infosys_list = []
#
#
# for each_row in range(1,max_rows_infosys+1):
#     infosys_list.append(str(infosys_sheet.cell(row = each_row,column = 1).value))
#
#
# #print infosys_list
# #----------------------------------------------------------------------------------------------------------------
#
# accenture_file = openpyxl.load_workbook('accenture.xlsx')
#
#
# accenture_sheet = accenture_file.get_sheet_by_name('Sheet1')
#
#
# max_rows_accenture = accenture_sheet.max_row
#
#
# accenture_list = []
#
#
# for each_row in range(1,max_rows_accenture+1):
#     accenture_list.append(str(accenture_sheet.cell(row = each_row,column = 1).value))
#
#
# #print accenture_list
# #---------------------------------------------------------------------------------------------------------------
#
# wipro_file = openpyxl.load_workbook('wipro.xlsx')
#
#
# wipro_sheet = wipro_file.get_sheet_by_name('Sheet1')
#
#
# max_rows_wipro = wipro_sheet.max_row
#
#
# wipro_list = []
#
#
# for each_row in range(1,max_rows_wipro+1):
#     wipro_list.append(str(wipro_sheet.cell(row = each_row,column = 1).value))
#
#
# print wipro_list
#---------------------------------------------------------------------------------------------------------------
infosys_list = read_compant_list('infosys.xlsx')
accenture_list = read_compant_list('accenture.xlsx')
wipro_list = read_compant_list('wipro.xlsx')
selected_students = {'INFOSYS':infosys_list,'ACCENTURE':accenture_list,'WIPRO':wipro_list}
#---------------------------------------------------------------------------------------------------------------
for each_student in options:
    try:
        preferences = options[each_student]
        flag = 0
        for each_preference in preferences:
            choice = preferences[each_preference]
            if each_student in selected_students[choice]:
                update_company_file(choice,each_student)
                flag = 1
                break
        if flag == 0:
            update_company_file('unplaced',each_student)
    except:
        print each_student,choice
#---------------------------------------------------------------------------------------------------------------
Infosys_placed.close()
Accenture_placed.close()
Wipro_placed.close()
#---------------------------------------------------------------------------------------------------------------
